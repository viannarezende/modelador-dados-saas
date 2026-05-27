import os
from pathlib import Path
from app.services.historico_service import registrar_historico_execucao


from dotenv import load_dotenv
from openai import OpenAI, RateLimitError
from openpyxl import load_workbook
from app.database.models import HistoricoExecucao
from app.database.connection import SessionLocal
from app.services.limites_service import (
    calcular_saldo,
    obter_ou_criar_uso_mensal,
    obter_plano_usuario,
    registrar_consumo,
    validar_limite,
)

BASE_DIR = Path(__file__).resolve().parents[2]
load_dotenv(dotenv_path=BASE_DIR / ".env", override=True)

MODEL_NAME = os.getenv("OPENAI_MODEL", "gpt-5.4")


def get_openai_client():
    api_key = os.getenv("OPENAI_API_KEY")

    if not api_key:
        raise ValueError("OPENAI_API_KEY não encontrada no ambiente.")

    return OpenAI(api_key=api_key)

def ler_arquivo_excel(caminho: str) -> str:
    path = Path(caminho)

    workbook = load_workbook(path, data_only=True)
    conteudo = []

    for sheet in workbook.worksheets:
        conteudo.append(f"\n=== Aba: {sheet.title} ===")

        rows = list(sheet.iter_rows(values_only=True))

        if not rows:
            continue

        headers = [str(h) if h else "" for h in rows[0]]

        for row in rows[1:]:
            linha_formatada = []

            for i, valor in enumerate(row):
                if valor is not None and i < len(headers):
                    chave = headers[i]
                    linha_formatada.append(f"{chave}: {valor}")

            if linha_formatada:
                conteudo.append(" | ".join(linha_formatada))

    return "\n".join(conteudo)
def ler_arquivo_excel(caminho: str) -> str:
    path = Path(caminho)

    workbook = load_workbook(path, data_only=True)
    conteudo = []

    for sheet in workbook.worksheets:
        conteudo.append(f"Aba: {sheet.title}")

        for row in sheet.iter_rows(values_only=True):
            valores = [str(valor) for valor in row if valor is not None]

            if valores:
                conteudo.append(" | ".join(valores))

    return "\n".join(conteudo)    

def ler_arquivo_texto(caminho: str | None) -> str:
    if not caminho:
        return "Nenhum arquivo enviado."

    path = Path(caminho)

    if not path.exists():
        return "Arquivo enviado, mas não encontrado no servidor."

    extensao = path.suffix.lower()

    if extensao in {".txt", ".csv"}:
        try:
            return path.read_text(encoding="utf-8")
        except UnicodeDecodeError:
            return path.read_text(encoding="latin-1")

    if extensao == ".xlsx":
        return ler_arquivo_excel(caminho)

    return (
        "Arquivo enviado em formato ainda não suportado para leitura automática. "
        "Use .txt, .csv ou .xlsx."
    )
def montar_prompt_modelagem(
    banco: str,
    etapa: str,
    padrao_nomenclatura: str,
    padrao_abreviacao: str,
    descricao: str,
    arquivo_nomenclatura: str | None = None,
    arquivo_abreviacao: str | None = None,
    modelo_anterior: str = ""
) -> str:
    
    conteudo_nomenclatura = ler_arquivo_texto(arquivo_nomenclatura)
    conteudo_abreviacao = ler_arquivo_texto(arquivo_abreviacao)
    return f"""
Você é um especialista em modelagem de dados.

Objetivo:
Gerar uma resposta técnica de modelagem de dados com base nas informações abaixo.

Banco de dados: {banco}
Etapa: {etapa}
Instrução específica da etapa:
- Se a etapa for "conceitual", gere entidades de negócio, relacionamentos e visão de alto nível. Não gere DDL.
- Se a etapa for "logico", gere tabelas, campos, chaves primárias, chaves estrangeiras, relacionamentos e regras de normalização. Não foque em sintaxe específica de banco. Não gere DDL.
- Se a etapa for "fisico", gere DDL compatível com o banco escolhido, tipos de dados, constraints, índices e observações técnicas.

Regras da etapa:
- A resposta DEVE respeitar a etapa selecionada.
- Não misture entregáveis de etapas diferentes.
- Use o banco selecionado para adaptar sintaxe, tipos de dados e recomendações.

Padrão de nomenclatura: {padrao_nomenclatura}
Padrão de abreviação: {padrao_abreviacao}

Descrição da solicitação:
{descricao}

Modelo anterior:
{modelo_anterior}

Arquivos auxiliares:
Conteúdo do arquivo de padrão de nomenclatura:
{conteudo_nomenclatura}

Conteúdo do arquivo de padrão de abreviação:
{conteudo_abreviacao}



Instruções:
- Responda em português do Brasil.
- Seja objetivo, técnico e organizado.
- Estruture a resposta para exibição em tela.
- Quando fizer sentido, apresente entidades, campos, relacionamentos e observações.
- Retorne a resposta em HTML puro (NÃO usar markdown).
- NÃO use símbolos como ###, **, etc.
- O modelo gerado DEVE obedecer rigorosamente ao conteúdo do arquivo de padrão de nomenclatura, quando enviado.
- O modelo gerado DEVE usar as abreviações conforme o arquivo de abreviação, quando enviado.
- Se houver conflito entre a descrição do usuário e os arquivos anexados, priorize os arquivos anexados.
- Não invente abreviações se o arquivo de abreviações trouxer uma regra específica.
- Ao criar nomes de tabelas, colunas, constraints e índices, aplique o padrão anexado.
- O modelo deve respeitar as 3 formas normais de banco de dados
- Aplicar até a Terceira Forma Normal (3FN):
- 1FN: garantir atomicidade dos atributos e eliminar grupos repetidos
- 2FN: eliminar dependências parciais da chave primária
- 3FN: eliminar dependências transitivas entre atributos
- Caso não tenha padrão de nomenclatura, utilize este abaixo:
- Nomes em português, maiúsculo, underscore
- Tabelas com prefixo (TB - tabelas padrão, RL-Tabelas relacionamento,HT- tabelas histórico,LG -Tabelas de Log etc.)
- Colunas com prefixo (ID, NM, DT, TP, VL)
- Evitar nulos
- Criar PK, FK, UK, CK
- Se a ação for ajuste, você deve modificar o modelo anterior, não criar um modelo do zero.
- Preserve tudo que não foi solicitado alterar.
- Aplique apenas as mudanças solicitadas pelo usuário.


- Utilize as seguintes tags HTML:
  <h2>, <h3>, <p>, <ul>, <li>, <strong>, <pre>, <code>

- Estruture a resposta da seguinte forma:

<h2>Título do modelo</h2>

<h3>Entidades</h3>
<ul>
  <li>...</li>
</ul>

<h3>Descrição</h3>
<p>...</p>

<h3>DDL</h3>
<pre>...</pre>

<h3>Observações</h3>
<p>...</p>
- Quando listar campos de tabelas, utilize tabelas HTML (<table>).
- Sempre que possível, apresente os campos no formato:

<table>
  <thead>
    <tr>
      <th>Campo</th>
      <th>Tipo</th>
      <th>Obrigatório</th>
      <th>Descrição</th>
    </tr>
  </thead>
  <tbody>
    ...
  </tbody>
</table>
""".strip()


def executar_modelagem(
    usuario_id: int,
    acao: str,
    descricao: str,
    banco: str,
    etapa: str,
    padrao_nomenclatura: str,
    padrao_abreviacao: str,
    arquivo_nomenclatura: str | None = None,
    arquivo_abreviacao: str | None = None,
    historico_origem_id: int | None = None,
) -> dict:
    db = SessionLocal()
    historico_origem = None
    if acao == "ajuste" and not historico_origem_id:
        return {
            "sucesso": False,
            "output": "Ajustes precisam estar vinculados a um modelo existente.",
            "tokens_entrada": 0,
            "tokens_saida": 0,
            "tokens_total": 0,
        }
    if acao == "ajuste":
        historico_origem = db.query(HistoricoExecucao).filter(
            HistoricoExecucao.id == historico_origem_id,
            HistoricoExecucao.usuario_id == usuario_id
        ).first()

    if acao == "ajuste" and not historico_origem:
     return {
        "sucesso": False,
        "output": "Modelo original não encontrado para ajuste.",
        "tokens_entrada": 0,
        "tokens_saida": 0,
        "tokens_total": 0,
    }
    try:
        plano = obter_plano_usuario(db, usuario_id)
        uso = obter_ou_criar_uso_mensal(db, usuario_id)

        permitido, mensagem = validar_limite(uso, plano, acao)
        if not permitido:
            saldo = calcular_saldo(plano, uso)
            return {
                "sucesso": False,
                "output": mensagem,
                "saldo": saldo,
                "tokens_entrada": 0,
                "tokens_saida": 0,
                "tokens_total": 0,
            }
        modelo_anterior = ""

        if acao == "ajuste" and historico_origem:
            modelo_anterior = historico_origem.resposta or ""

        prompt = montar_prompt_modelagem(
            banco=banco,
            etapa=etapa,
            padrao_nomenclatura=padrao_nomenclatura,
            padrao_abreviacao=padrao_abreviacao,
            descricao=descricao,
            arquivo_nomenclatura=arquivo_nomenclatura,
            arquivo_abreviacao=arquivo_abreviacao,
            modelo_anterior=modelo_anterior,
        )

        client = get_openai_client()

        response = client.responses.create(
            model=MODEL_NAME,
            input=prompt,
        )

        output_text = getattr(response, "output_text", None) or "Nenhuma resposta retornada."

        usage = getattr(response, "usage", None)
        tokens_entrada = getattr(usage, "input_tokens", 0) if usage else 0
        tokens_saida = getattr(usage, "output_tokens", 0) if usage else 0
        tokens_total = getattr(usage, "total_tokens", tokens_entrada + tokens_saida) if usage else 0

        uso = registrar_consumo(
            db=db,
            uso=uso,
            acao=acao,
            tokens_entrada=tokens_entrada,
            tokens_saida=tokens_saida,
        )

        saldo = calcular_saldo(plano, uso)
        registrar_historico_execucao(
            db=db,
            usuario_id=usuario_id,
            acao=acao,
            banco=banco,
            etapa=etapa,
            descricao=descricao,
            padrao_nomenclatura=padrao_nomenclatura,
            padrao_abreviacao=padrao_abreviacao,
            arquivo_nomenclatura=arquivo_nomenclatura,
            arquivo_abreviacao=arquivo_abreviacao,
            historico_origem_id=historico_origem_id,
            status="sucesso",
            resposta=output_text,
            tokens_entrada=tokens_entrada,
            tokens_saida=tokens_saida,
            tokens_total=tokens_total,
        )

        return {
            "sucesso": True,
            "output": output_text,
            "saldo": saldo,
            "tokens_entrada": tokens_entrada,
            "tokens_saida": tokens_saida,
            "tokens_total": tokens_total,
        }

    except RateLimitError as exc:
        mensagem = str(exc)

        if "insufficient_quota" in mensagem or "exceeded your current quota" in mensagem:
            mensagem = (
                "A integração com a OpenAI está configurada, mas a conta está sem saldo/cota no momento. "
                "Assim que a cota for liberada, a geração voltará a funcionar normalmente."
            )
        else:
            mensagem = f"Limite temporário da API atingido: {str(exc)}"

        saldo = calcular_saldo(plano, uso)
        registrar_historico_execucao(
            db=db,
            usuario_id=usuario_id,
            acao=acao,
            banco=banco,
            etapa=etapa,
            descricao=descricao,
            padrao_nomenclatura=padrao_nomenclatura,
            padrao_abreviacao=padrao_abreviacao,
            arquivo_nomenclatura=arquivo_nomenclatura,
            arquivo_abreviacao=arquivo_abreviacao,
            historico_origem_id=historico_origem_id,
            status="erro",
            resposta=mensagem,
        )

        return {
            "sucesso": False,
            "output": mensagem,
            "saldo": saldo,
            "tokens_entrada": 0,
            "tokens_saida": 0,
            "tokens_total": 0,
        }   

    except Exception as exc:
        saldo = calcular_saldo(plano, uso)
        registrar_historico_execucao(
            db=db,
            usuario_id=usuario_id,
            acao=acao,
            banco=banco,
            etapa=etapa,
            descricao=descricao,
            padrao_nomenclatura=padrao_nomenclatura,
            padrao_abreviacao=padrao_abreviacao,
            arquivo_nomenclatura=arquivo_nomenclatura,
            arquivo_abreviacao=arquivo_abreviacao,
            historico_origem_id=historico_origem_id,
            status="erro",
            resposta=str(exc),
        )

        return {
            "sucesso": False,
            "output": f"Erro ao executar modelagem: {str(exc)}",
            "saldo": saldo,
            "tokens_entrada": 0,
            "tokens_saida": 0,
            "tokens_total": 0,
        }

    finally:
        db.close()